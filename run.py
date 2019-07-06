import os
import requests
import arrow
import time
from collections import defaultdict
from openpyxl import load_workbook


API_TOKEN = os.environ['TOGGL_API_TOKEN']
WORKSPACE_ID = os.environ['TOGGL_WORKSPACE_ID']
EMAIL = os.environ['TOGGL_EMAIL']
EXCEL_FILE = os.environ['TOGGL_EXCEL_FILE']
SHEET_NAME = os.environ['TOGGL_EXCEL_SHEET_NAME']


def get_toggl(page=1):
    headers = {'content-type': 'application/json'}
    params = {
        'page': page,
        'user_agent': EMAIL,
        'workspace_id': WORKSPACE_ID,
        'since': '2019-03-01',
        'until': str(arrow.now().date()),
    }

    auth = requests.auth.HTTPBasicAuth(API_TOKEN, 'api_token')

    return requests.get(
        'https://toggl.com/reports/api/v2/details',
        auth=auth, headers=headers, params=params
    )


if __name__ == '__main__':
    page = 1
    summary = get_toggl(page).json()
    get_next = True
    data_by_sim = defaultdict(int)
    book = load_workbook(EXCEL_FILE)
    sheet = book.get_sheet_by_name(SHEET_NAME)

    # clear excel
    for a in sheet['A':'B']:
        for cell in a:
            cell.value = None

    while get_next:
        summary = get_toggl(page).json()

        for data in summary['data']:
            description = data['description']
            start = arrow.get(data['start']).timestamp
            end = arrow.get(data['end']).timestamp
            diff = end - start

            if '[' not in description or ']' not in description:
                continue

            description = description.split(']')[0].replace('[', '').strip()
            data_by_sim[description] += end - start

        # pagination logic
        page += 1
        get_next = len(summary['data']) == summary['per_page']
        if get_next:
            time.sleep(1)

    row = 0
    for key, value in sorted(data_by_sim.items(), key=lambda x: x[0]):
        row += 1
        m, s = divmod(value, 60)
        h, m = divmod(m, 60)
        total_time = '{:02d}:{:02d}:{:02d}'.format(h, m, s)

        sheet.cell(row=row, column=1).value = key
        sheet.cell(row=row, column=2).value = total_time

    # save excel
    book.save(EXCEL_FILE)
